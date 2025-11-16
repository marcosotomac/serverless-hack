"""
Handler para exportar incidentes a PDF, Excel o CSV
Solo accesible para rol 'autoridad'
"""
import os
import json
import io
import csv
from datetime import datetime
from decimal import Decimal
import boto3
from src.common.auth import authorize
from src.common.response import json_response
from src.common.dynamodb import list_all_incidents


s3_client = boto3.client("s3")


class DecimalEncoder(json.JSONEncoder):
    """Encoder personalizado para Decimal"""

    def default(self, obj):
        if isinstance(obj, Decimal):
            return int(obj) if obj % 1 == 0 else float(obj)
        if isinstance(obj, set):
            return list(obj)
        return super(DecimalEncoder, self).default(obj)


def export_to_csv(incidents):
    """
    Exporta incidentes a formato CSV
    """
    output = io.StringIO()

    if not incidents:
        return ""

    # Definir campos a exportar
    fields = [
        "incidentId", "type", "location", "description",
        "urgency", "priority", "status", "reportedBy",
        "reporterRole", "assignedTo", "createdAt", "updatedAt",
        "significanceCount"
    ]

    writer = csv.DictWriter(output, fieldnames=fields, extrasaction='ignore')
    writer.writeheader()

    for incident in incidents:
        # Convertir sets a strings
        row = {}
        for field in fields:
            value = incident.get(field, "")
            if isinstance(value, set):
                value = ", ".join(str(v) for v in value)
            elif isinstance(value, Decimal):
                value = int(value) if value % 1 == 0 else float(value)
            row[field] = value
        writer.writerow(row)

    return output.getvalue()


def export_to_excel(incidents):
    """
    Exporta incidentes a formato Excel (usando openpyxl)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise Exception(
            "openpyxl no está instalado. Instale con: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "Incidentes"

    # Headers
    headers = [
        "ID", "Tipo", "Ubicación", "Descripción",
        "Urgencia", "Prioridad", "Estado", "Reportado Por",
        "Rol Reportador", "Asignado A", "Creado", "Actualizado",
        "Votos Significancia"
    ]

    # Estilo para headers
    header_fill = PatternFill(start_color="366092",
                              end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Datos
    for row_idx, incident in enumerate(incidents, start=2):
        ws.cell(row=row_idx, column=1, value=incident.get("incidentId", ""))
        ws.cell(row=row_idx, column=2, value=incident.get("type", ""))
        ws.cell(row=row_idx, column=3, value=incident.get("location", ""))
        ws.cell(row=row_idx, column=4, value=incident.get("description", ""))
        ws.cell(row=row_idx, column=5, value=incident.get("urgency", ""))
        ws.cell(row=row_idx, column=6, value=incident.get("priority", ""))
        ws.cell(row=row_idx, column=7, value=incident.get("status", ""))
        ws.cell(row=row_idx, column=8, value=incident.get("reportedBy", ""))
        ws.cell(row=row_idx, column=9, value=incident.get("reporterRole", ""))
        ws.cell(row=row_idx, column=10, value=incident.get("assignedTo", ""))
        ws.cell(row=row_idx, column=11, value=incident.get("createdAt", ""))
        ws.cell(row=row_idx, column=12, value=incident.get("updatedAt", ""))

        sig_count = incident.get("significanceCount", 0)
        if isinstance(sig_count, Decimal):
            sig_count = int(sig_count)
        ws.cell(row=row_idx, column=13, value=sig_count)

    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Columna de descripción más ancha
    ws.column_dimensions['D'].width = 40

    # Guardar en memoria
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.getvalue()


def export_to_pdf(incidents):
    """
    Exporta incidentes a formato PDF (usando reportlab)
    """
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
    except ImportError:
        raise Exception(
            "reportlab no está instalado. Instale con: pip install reportlab")

    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    elements = []

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1a365d'),
        spaceAfter=30,
        alignment=TA_CENTER
    )

    # Título
    title = Paragraph("Reporte de Incidentes", title_style)
    elements.append(title)

    # Fecha de generación
    date_style = ParagraphStyle(
        'DateStyle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        alignment=TA_CENTER,
        spaceAfter=20
    )
    date_text = Paragraph(
        f"Generado: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
        date_style
    )
    elements.append(date_text)
    elements.append(Spacer(1, 0.5*inch))

    # Resumen
    summary_data = {
        "total": len(incidents),
        "pendientes": sum(1 for i in incidents if i.get("status") == "pendiente"),
        "en_atencion": sum(1 for i in incidents if i.get("status") == "en_atencion"),
        "resueltos": sum(1 for i in incidents if i.get("status") == "resuelto"),
    }

    summary_table = Table([
        ["Total", "Pendientes", "En Atención", "Resueltos"],
        [summary_data["total"], summary_data["pendientes"],
         summary_data["en_atencion"], summary_data["resueltos"]]
    ])

    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))

    elements.append(summary_table)
    elements.append(Spacer(1, 0.5*inch))

    # Detalles de incidentes (primeros 50 para no exceder tamaño)
    if incidents:
        elements.append(
            Paragraph("Detalles de Incidentes", styles['Heading2']))
        elements.append(Spacer(1, 0.2*inch))

        for idx, incident in enumerate(incidents[:50], start=1):
            incident_data = [
                ["ID:", incident.get("incidentId", "N/A")],
                ["Tipo:", incident.get("type", "N/A")],
                ["Ubicación:", incident.get("location", "N/A")],
                ["Estado:", incident.get("status", "N/A")],
                ["Urgencia:", incident.get("urgency", "N/A")],
                ["Reportado por:", incident.get("reportedBy", "N/A")],
                ["Asignado a:", incident.get("assignedTo", "No asignado")],
                ["Creado:", incident.get("createdAt", "N/A")[:10]],
            ]

            incident_table = Table(incident_data, colWidths=[1.5*inch, 4*inch])
            incident_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#e2e8f0')),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 9),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))

            elements.append(incident_table)
            elements.append(Spacer(1, 0.3*inch))

            # Page break cada 5 incidentes
            if idx % 5 == 0 and idx < len(incidents[:50]):
                elements.append(PageBreak())

    if len(incidents) > 50:
        note = Paragraph(
            f"<i>Nota: Se muestran los primeros 50 de {len(incidents)} incidentes totales.</i>",
            styles['Normal']
        )
        elements.append(note)

    doc.build(elements)
    output.seek(0)

    return output.getvalue()


@authorize(["autoridad"])
def handler(event, context):
    """
    Exporta incidentes en el formato solicitado
    POST /analytics/export
    Body: { format: "csv" | "excel" | "pdf", filters?: {...} }
    """
    try:
        body = json.loads(event.get("body", "{}"))
    except json.JSONDecodeError:
        return json_response(400, {"error": "JSON inválido"})

    export_format = body.get("format", "csv").lower()
    filters = body.get("filters", {})

    if export_format not in ["csv", "excel", "pdf"]:
        return json_response(400, {
            "error": "Formato no soportado. Use: csv, excel o pdf"
        })

    try:
        # Obtener todos los incidentes
        incidents = list_all_incidents()

        # Aplicar filtros si existen
        if filters:
            if "status" in filters:
                incidents = [i for i in incidents if i.get(
                    "status") == filters["status"]]
            if "type" in filters:
                incidents = [i for i in incidents if i.get(
                    "type") == filters["type"]]
            if "urgency" in filters:
                incidents = [i for i in incidents if i.get(
                    "urgency") == filters["urgency"]]

        # Ordenar por fecha (más recientes primero)
        incidents.sort(key=lambda x: x.get("createdAt", ""), reverse=True)

        # Generar archivo según formato
        if export_format == "csv":
            content = export_to_csv(incidents)
            content_type = "text/csv"
            filename = f"incidentes_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
            content_bytes = content.encode('utf-8')

        elif export_format == "excel":
            content_bytes = export_to_excel(incidents)
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            filename = f"incidentes_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"

        else:  # pdf
            content_bytes = export_to_pdf(incidents)
            content_type = "application/pdf"
            filename = f"incidentes_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.pdf"

        # Subir a S3
        results_bucket = os.environ["ANALYTICS_RESULTS_BUCKET"]
        s3_key = f"exports/{filename}"

        s3_client.put_object(
            Bucket=results_bucket,
            Key=s3_key,
            Body=content_bytes,
            ContentType=content_type,
            ContentDisposition=f'attachment; filename="{filename}"'
        )

        # Generar URL prefirmada para descarga (válida por 1 hora)
        download_url = s3_client.generate_presigned_url(
            'get_object',
            Params={'Bucket': results_bucket, 'Key': s3_key},
            ExpiresIn=3600
        )

        return json_response(200, {
            "message": "Exportación completada",
            "format": export_format,
            "filename": filename,
            "incidentsCount": len(incidents),
            "downloadUrl": download_url,
            "expiresIn": "1 hora"
        })

    except Exception as e:
        return json_response(500, {
            "error": f"Error en exportación: {str(e)}"
        })
